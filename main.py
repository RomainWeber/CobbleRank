import json
import os
import pandas as pd
import numpy as np
import configparser
import openpyxl
import datetime
import ftplib
import math
import warnings
import paramiko
import excel2img



def loadCobblemonData(csvtoggle, csvpath, inputmode, ftpserver, ftppath):
    df = pd.DataFrame()
    root_dirnames = []
    if inputmode == "ftp" or inputmode == "sftp":
        if ftppath == "":
            ftppath_complete = "world/cobblemonplayerdata"
        else:
            ftppath_complete = ftppath + "/world/cobblemonplayerdata"
        if inputmode == "ftp":
            ftpserver.cwd(ftppath)
            with open("data/usercache/usercache.json", "wb") as file:
                ftpserver.retrbinary(f"RETR usercache.json", file.write)
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            root_dirnames = ftpserver.nlst(ftppath_complete)
            ftpserver.cwd(ftppath_complete)
        else:
            ftpserver.chdir(ftppath)
            ftpserver.get("usercache.json", "data/usercache/usercache.json")
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            # Go back to root
            # Obtenir le répertoire courant
            current_dir = ftp_server.normalize(".")

            # Calculer combien de niveaux remonter
            num_levels = len(current_dir.split("/")) - 1

            # Remonter jusqu'à la racine
            ftp_server.chdir("/" if num_levels == 0 else "../" * num_levels)
            # Get directories
            root_dirnames = ftpserver.listdir(ftppath_complete)
            ftpserver.chdir(ftppath_complete)
            
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            # Go to the subfolder
            if inputmode == "ftp":
                ftpserver.cwd(dirname.split("/")[-1])
                filenames = ftpserver.nlst()
            else:
                ftpserver.chdir(dirname.split("/")[-1])
                filenames = ftpserver.listdir()
            
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                # Download the file to process
                local_file = "data/cobblemonplayerdata/"+filename
                with open(local_file, "wb") as file:
                    if inputmode == "ftp":
                        ftpserver.retrbinary(f"RETR {filename}", file.write)
                    else:
                        ftpserver.get(filename, local_file)
                
                with open(local_file, "r") as file:
                    data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                
            if inputmode == "ftp":
                ftpserver.cwd("../")  # Move back to the parent directory
            else:
                ftpserver.chdir("..")
        # Go back to root
        if inputmode == "ftp":
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
        else:
            # Obtenir le répertoire courant
            current_dir = ftp_server.normalize(".")

            # Calculer combien de niveaux remonter
            num_levels = len(current_dir.split("/")) - 1

            # Remonter jusqu'à la racine
            ftp_server.chdir("/" if num_levels == 0 else "../" * num_levels)
    else:
        names_file = open('data/usercache/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        i = -1
        path = 'data/cobblemonplayerdata'
        for dirpath, dirnames, filenames in os.walk(path):
            if len(dirnames) > 0:
                root_dirnames = dirnames
            for filename in filenames:
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                file = open(path + '/' + root_dirnames[i] + '/' + filename)
                data = json.load(file)['extraData']['cobbledex_discovery']['registers']
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                temp_df = temp_df.transpose().iloc[:]
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
            i += 1
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df

def most_pokemons_leaderboard(df, config, type):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    
    if type == "standard":
        sheet_name = "leaderboard2"
    elif type == "shiny":
        sheet_name = "leaderboard3"
    elif type == "legendary":
        sheet_name = "leaderboard4"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['COBBLEMONLEADERBOARDS']['ExcelRows'])
    ExcelCols = int(config['COBBLEMONLEADERBOARDS']['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated']))
    ws.cell(row=ExcelRows+4, column=2, value=config['COBBLEMONLEADERBOARDS']['Subtitle'])
    wb.save(file_path)

# Read config
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf8')

# Connect to FTP if activated
ftp_server = None
if config['INPUT']['Mode'] == "ftp":
    ftp_server = ftplib.FTP(config['INPUT']['Host'], open("username.txt", "r").read(), open("password.txt", "r").read())
    ftp_server.encoding = "utf-8"
if config['INPUT']['Mode'] == "sftp":
    transport = paramiko.Transport((config['INPUT']['Host'], int(config['INPUT']['Port'])))
    transport.connect(username="theskyrimlegend1.4a5ca669", password="3277Sun232Aug-MSR")
    ftp_server = paramiko.SFTPClient.from_transport(transport)

# Load the data
print("LOADING COBBLEMON DATA")
if config['GLOBALMATRIX']['UseCSV'] == "false":
    cobblemon_df = loadCobblemonData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['INPUT']['Mode'], ftp_server, config['INPUT']['FTPPath'])
else:
    cobblemon_df = pd.read_csv(config['GLOBALMATRIX']['CSVPath'], index_col=[0,1,2], skipinitialspace=True)

# Close the Connection
if config['INPUT']['Mode'] == "ftp":
    ftp_server.quit()
if config['INPUT']['Mode'] == "sftp":
    ftp_server.close()

# Prepare the counting DF
count_df = cobblemon_df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
pokemons_db = pd.read_csv('Pokemon.csv')
legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

# Total leaderboard feature
if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config, "standard")
    excel2img.export_img("output.xlsx", "result/pokedex.png", "leaderboard2","A1:N15")

# Shiny leaderboard feature
if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":
    player_sum = pd.DataFrame(((cobblemon_df == "True") | (cobblemon_df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config, "shiny")
    excel2img.export_img("output.xlsx", "result/shiny.png", "leaderboard3","A1:N15")

# Legendary leaderboard feature
if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
    legs = legendary_list['Cobblemon'].tolist()
    leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)
    #leg_count_df.to_csv("temp.csv")
    player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    most_pokemons_leaderboard(player_sum, config, "legendary")
    excel2img.export_img("output.xlsx", "result/legendary.png", "leaderboard4","A1:N15")

print("Done!")